#!/usr/bin/env python3
"""
RFQ Supplier Finder Script
Processes an RFQ file (CSV or Excel), searches for suppliers for each part number,
and generates a report with pricing and ordering information.

Integrates with KIP workflow using Google Custom Search API and OpenAI.

Required Environment Variables:
- GOOGLE_API_KEY: Your Google Custom Search API key
- GOOGLE_CX: Your Google Custom Search Engine ID
- OPENAI_API_KEY: OpenAI API key for intelligent data extraction

Usage:
    python rfq_supplier_finder.py input_rfq.csv [--output results.txt] [--delay 1.0]
    python rfq_supplier_finder.py input_rfq.xlsx [--output results.txt] [--delay 1.0]

Input Format (flexible - will detect columns):
    Part Number, Description, Quantity, Unit (optional columns)

Example:
    python rfq_supplier_finder.py parts_list.xlsx --output supplier_quotes.txt
"""

import os
import sys
import csv
import json
import time
import argparse
import re
from datetime import datetime
from typing import List, Dict, Optional, Any
from pathlib import Path
import requests
from openai import OpenAI
from dotenv import load_dotenv

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

# Load environment variables
load_dotenv()

# Configuration
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
GOOGLE_CX = os.getenv("GOOGLE_CX")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# Search settings
MAX_SEARCH_RESULTS = 10
REQUEST_DELAY = 1.0  # seconds between requests


class SupplierFinder:
    """Finds suppliers and pricing for parts using web search and AI extraction."""

    def __init__(self, delay: float = 1.0):
        self.delay = delay
        self.openai_client = OpenAI(api_key=OPENAI_API_KEY)
        self._validate_config()

    def _validate_config(self):
        """Validate required API keys are present."""
        missing = []
        if not GOOGLE_API_KEY:
            missing.append("GOOGLE_API_KEY")
        if not GOOGLE_CX:
            missing.append("GOOGLE_CX")
        if not OPENAI_API_KEY:
            missing.append("OPENAI_API_KEY")

        if missing:
            print(f"ERROR: Missing required environment variables: {', '.join(missing)}")
            print("Please set these in your .env file:")
            print("  GOOGLE_API_KEY=your_google_api_key")
            print("  GOOGLE_CX=your_custom_search_engine_id")
            print("  OPENAI_API_KEY=your_openai_api_key")
            sys.exit(1)

    def google_search(self, query: str) -> List[Dict]:
        """
        Perform a Google Custom Search for the given query.
        Returns list of search results with title, link, and snippet.
        """
        url = "https://www.googleapis.com/customsearch/v1"
        params = {
            "key": GOOGLE_API_KEY,
            "cx": GOOGLE_CX,
            "q": query,
            "num": MAX_SEARCH_RESULTS
        }

        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()

            results = []
            for item in data.get("items", []):
                results.append({
                    "title": item.get("title", ""),
                    "link": item.get("link", ""),
                    "snippet": item.get("snippet", ""),
                    "displayLink": item.get("displayLink", "")
                })
            return results

        except requests.exceptions.RequestException as e:
            print(f"  Search error: {e}")
            return []

    def fetch_page_content(self, url: str, max_chars: int = 15000) -> Optional[str]:
        """Fetch and extract text content from a webpage."""
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }
            response = requests.get(url, headers=headers, timeout=15)
            response.raise_for_status()

            # Simple HTML text extraction
            from html.parser import HTMLParser

            class TextExtractor(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.text_parts = []
                    self.skip_tags = {'script', 'style', 'nav', 'header', 'footer'}
                    self.current_tag = None

                def handle_starttag(self, tag, attrs):
                    self.current_tag = tag

                def handle_endtag(self, tag):
                    self.current_tag = None

                def handle_data(self, data):
                    if self.current_tag not in self.skip_tags:
                        text = data.strip()
                        if text:
                            self.text_parts.append(text)

            extractor = TextExtractor()
            extractor.feed(response.text)
            content = ' '.join(extractor.text_parts)

            # Truncate if too long
            if len(content) > max_chars:
                content = content[:max_chars] + "..."

            return content

        except Exception as e:
            print(f"  Failed to fetch {url}: {e}")
            return None

    def extract_supplier_info(self, part_number: str, part_description: str,
                              quantity: int, search_results: List[Dict],
                              page_contents: List[str]) -> Dict:
        """
        Use OpenAI to extract supplier information, pricing, and ordering details
        from search results and page content.
        """
        # Build context from search results and page contents
        context_parts = []
        for i, result in enumerate(search_results):
            context_parts.append(f"Source {i + 1}: {result['title']}")
            context_parts.append(f"URL: {result['link']}")
            context_parts.append(f"Preview: {result['snippet']}")
            if i < len(page_contents) and page_contents[i]:
                # Include first 3000 chars of each page
                context_parts.append(f"Page Content: {page_contents[i][:3000]}")
            context_parts.append("---")

        context = "\n".join(context_parts)

        prompt = f"""You are an expert procurement specialist analyzing search results to find suppliers for a specific part.

CRITICAL REQUIREMENTS:
- The part number must match EXACTLY: "{part_number}"
- Do NOT substitute with similar or compatible parts
- If you cannot find the exact part number, say "EXACT MATCH NOT FOUND"
- Only report suppliers that explicitly list or sell this exact part number

PART INFORMATION:
- Part Number: {part_number}
- Description: {part_description or "Not provided"}
- Quantity Needed: {quantity}

SEARCH RESULTS AND PAGE CONTENT:
{context}

Based on the above information, extract supplier details for this EXACT part number.
Respond in JSON format:

{{
    "exact_match_found": true/false,
    "suppliers": [
        {{
            "supplier_name": "Company name",
            "website": "URL",
            "contact_info": {{
                "phone": "phone number or null",
                "email": "email or null",
                "address": "address or null"
            }},
            "pricing": {{
                "unit_price": "price per unit or 'Quote Required' or 'Not Listed'",
                "currency": "USD",
                "minimum_order_qty": "MOQ if listed",
                "lead_time": "Lead time if listed",
                "price_notes": "Any pricing conditions or notes"
            }},
            "ordering_process": "Step-by-step process to order from this supplier",
            "confidence": "HIGH/MEDIUM/LOW - how confident you are this supplier has the exact part",
            "notes": "Any additional relevant information"
        }}
    ],
    "alternative_sources": "If exact match not found, suggest where to look (distributors, OEM, etc.)",
    "search_recommendations": "Additional search terms that might help find this part"
}}

If no suppliers are found, return empty suppliers array and provide helpful recommendations."""

        try:
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system",
                     "content": "You are a procurement specialist. Extract accurate supplier information. Never invent or hallucinate data - only report what you can verify from the provided content."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )

            result = json.loads(response.choices[0].message.content)
            return result

        except Exception as e:
            print(f"  AI extraction error: {e}")
            return {
                "exact_match_found": False,
                "suppliers": [],
                "error": str(e),
                "alternative_sources": "Error occurred during analysis",
                "search_recommendations": "Try manual search"
            }

    def find_suppliers_for_part(self, part_number: str, description: str = "",
                                quantity: int = 1) -> Dict:
        """
        Find suppliers for a specific part number.
        Returns extracted supplier information.
        """
        print(f"\n  Searching for: {part_number}")

        # Build search queries - exact part number focused
        queries = [
            f'"{part_number}" supplier buy price',
            f'"{part_number}" distributor order',
            f'"{part_number}" purchase quote'
        ]

        # Add description-based query if available
        if description:
            queries.append(f'"{part_number}" {description[:50]}')

        all_results = []
        seen_urls = set()

        for query in queries[:3]:  # Limit to 3 queries per part
            results = self.google_search(query)
            for result in results:
                if result['link'] not in seen_urls:
                    seen_urls.add(result['link'])
                    all_results.append(result)
            time.sleep(self.delay)

        if not all_results:
            print(f"    No search results found")
            return {
                "part_number": part_number,
                "exact_match_found": False,
                "suppliers": [],
                "alternative_sources": "No search results. Try searching manufacturer directly.",
                "search_recommendations": "Contact OEM or authorized distributors"
            }

        print(f"    Found {len(all_results)} unique results, fetching pages...")

        # Fetch top page contents (limit to save time/API calls)
        page_contents = []
        for result in all_results[:5]:
            content = self.fetch_page_content(result['link'])
            page_contents.append(content)
            time.sleep(0.5)

        print(f"    Analyzing with AI...")

        # Extract supplier info using AI
        supplier_info = self.extract_supplier_info(
            part_number, description, quantity,
            all_results[:5], page_contents
        )

        supplier_info["part_number"] = part_number
        supplier_info["description"] = description
        supplier_info["quantity_requested"] = quantity
        supplier_info["sources_searched"] = len(all_results)

        return supplier_info


def read_rfq_file(filepath: str) -> List[Dict]:
    """
    Read an RFQ file (CSV or Excel) and extract part information.
    Flexible column detection - looks for common column names.
    """
    file_ext = Path(filepath).suffix.lower()

    if file_ext in ['.xlsx', '.xls', '.xlsm']:
        return read_rfq_excel(filepath)
    else:
        return read_rfq_csv(filepath)


def read_rfq_excel(filepath: str) -> List[Dict]:
    """Read an RFQ Excel file and extract part information."""

    # Try pandas first (better handling), fall back to openpyxl
    if HAS_PANDAS:
        return read_rfq_excel_pandas(filepath)
    elif HAS_OPENPYXL:
        return read_rfq_excel_openpyxl(filepath)
    else:
        print("ERROR: Excel file detected but neither pandas nor openpyxl is installed.")
        print("Install with: pip install pandas openpyxl")
        sys.exit(1)


def read_rfq_excel_pandas(filepath: str) -> List[Dict]:
    """Read Excel file using pandas."""
    import pandas as pd

    # Common column name variations (order matters - more specific first)
    part_number_cols = ['part no.', 'part no', 'part_number', 'part number', 'partnumber',
                        'pn', 'part#', 'part_no', 'partno', 'item_number', 'item number',
                        'sku', 'mpn', 'manufacturer_part_number', 'nsn', 'niin',
                        'item#', 'item no', 'p/n', 'mfr part', 'mfg part', 'part']
    description_cols = ['description', 'desc', 'item_description', 'part_description',
                        'name', 'item_name', 'part_name', 'title', 'nomenclature']
    quantity_cols = ['quantity', 'qty', 'amount', 'count', 'units', 'order_qty', 'req qty']
    unit_cols = ['unit', 'uom', 'unit_of_measure', 'each', 'u/m']

    def find_column(columns: List[str], possible_names: List[str]) -> Optional[str]:
        """Find column name by trying multiple possible names."""
        cols_lower = {str(c).lower().strip(): c for c in columns}
        # First try exact matches
        for name in possible_names:
            if name.lower() in cols_lower:
                return cols_lower[name.lower()]
        # Then try partial matches (column contains the search term)
        for name in possible_names:
            for col_lower, col_orig in cols_lower.items():
                if name.lower() in col_lower:
                    return col_orig
        return None

    # Read Excel file - try to find the header row
    df = pd.read_excel(filepath, header=None)

    # Find header row (look for row containing part number-like column)
    header_row = 0
    for idx, row in df.iterrows():
        row_values = [str(v).lower().strip() for v in row.values if pd.notna(v)]
        for val in row_values:
            if any(pn in val for pn in ['part', 'item', 'sku', 'p/n', 'nsn']):
                header_row = idx
                break
        if header_row > 0:
            break

    # Re-read with correct header
    df = pd.read_excel(filepath, header=header_row)

    # Clean column names
    df.columns = [str(c).strip() for c in df.columns]

    print(f"Detected columns: {list(df.columns)}")

    # Find column mappings
    pn_col = find_column(df.columns, part_number_cols)
    desc_col = find_column(df.columns, description_cols)
    qty_col = find_column(df.columns, quantity_cols)
    unit_col = find_column(df.columns, unit_cols)

    if pn_col is None:
        # If no part number column found, try first non-empty column
        for col in df.columns:
            if df[col].notna().any():
                pn_col = col
                print(f"Warning: No part number column detected, using '{col}'")
                break

    print(f"\nColumn mapping:")
    print(f"  Part Number: {pn_col}")
    print(f"  Description: {desc_col or 'Not found'}")
    print(f"  Quantity: {qty_col or 'Not found (defaulting to 1)'}")
    print(f"  Unit: {unit_col or 'Not found'}")

    parts = []
    for idx, row in df.iterrows():
        # Skip empty rows
        if pd.isna(row[pn_col]) or str(row[pn_col]).strip() == '':
            continue

        part_number = str(row[pn_col]).strip()

        # Skip header-like rows
        if part_number.lower() in part_number_cols:
            continue

        part = {
            "part_number": part_number,
            "description": "",
            "quantity": 1,
            "unit": ""
        }

        if desc_col and pd.notna(row.get(desc_col)):
            part["description"] = str(row[desc_col]).strip()

        if qty_col and pd.notna(row.get(qty_col)):
            try:
                qty_val = row[qty_col]
                if isinstance(qty_val, (int, float)):
                    part["quantity"] = int(qty_val)
                else:
                    qty_str = str(qty_val).strip().replace(',', '')
                    part["quantity"] = int(float(qty_str)) if qty_str else 1
            except (ValueError, TypeError):
                part["quantity"] = 1

        if unit_col and pd.notna(row.get(unit_col)):
            part["unit"] = str(row[unit_col]).strip()

        parts.append(part)

    return parts


def read_rfq_excel_openpyxl(filepath: str) -> List[Dict]:
    """Read Excel file using openpyxl."""
    from openpyxl import load_workbook

    # Common column name variations (order matters - more specific first)
    part_number_cols = ['part no.', 'part no', 'part_number', 'part number', 'partnumber',
                        'pn', 'part#', 'part_no', 'partno', 'item_number', 'item number',
                        'sku', 'mpn', 'manufacturer_part_number', 'nsn', 'niin',
                        'item#', 'item no', 'p/n', 'mfr part', 'mfg part', 'part']
    description_cols = ['description', 'desc', 'item_description', 'part_description',
                        'name', 'item_name', 'part_name', 'title', 'nomenclature']
    quantity_cols = ['quantity', 'qty', 'amount', 'count', 'units', 'order_qty', 'req qty']
    unit_cols = ['unit', 'uom', 'unit_of_measure', 'each', 'u/m']

    def find_column(headers: List[str], possible_names: List[str]) -> Optional[int]:
        """Find column index by trying multiple possible names."""
        headers_lower = [str(h).lower().strip() if h else '' for h in headers]
        # First try exact matches
        for name in possible_names:
            for idx, header in enumerate(headers_lower):
                if name.lower() == header:
                    return idx
        # Then try partial matches
        for name in possible_names:
            for idx, header in enumerate(headers_lower):
                if name.lower() in header:
                    return idx
        return None

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    # Find header row
    header_row_idx = 0
    for idx, row in enumerate(rows):
        row_values = [str(v).lower().strip() for v in row if v is not None]
        for val in row_values:
            if any(pn in val for pn in ['part', 'item', 'sku', 'p/n', 'nsn']):
                header_row_idx = idx
                break
        if header_row_idx > 0:
            break

    headers = [str(h) if h else '' for h in rows[header_row_idx]]
    print(f"Detected columns: {[h for h in headers if h]}")

    # Find column indices
    pn_idx = find_column(headers, part_number_cols)
    desc_idx = find_column(headers, description_cols)
    qty_idx = find_column(headers, quantity_cols)
    unit_idx = find_column(headers, unit_cols)

    if pn_idx is None:
        pn_idx = 0
        print(f"Warning: No part number column detected, using first column")

    print(f"\nColumn mapping:")
    print(f"  Part Number: {headers[pn_idx] if pn_idx < len(headers) else 'Column 0'}")
    print(f"  Description: {headers[desc_idx] if desc_idx and desc_idx < len(headers) else 'Not found'}")
    print(f"  Quantity: {headers[qty_idx] if qty_idx and qty_idx < len(headers) else 'Not found (defaulting to 1)'}")

    parts = []
    for row in rows[header_row_idx + 1:]:
        if not row or pn_idx >= len(row) or not row[pn_idx]:
            continue

        part_number = str(row[pn_idx]).strip()
        if not part_number or part_number.lower() in part_number_cols:
            continue

        part = {
            "part_number": part_number,
            "description": "",
            "quantity": 1,
            "unit": ""
        }

        if desc_idx is not None and desc_idx < len(row) and row[desc_idx]:
            part["description"] = str(row[desc_idx]).strip()

        if qty_idx is not None and qty_idx < len(row) and row[qty_idx]:
            try:
                qty_val = row[qty_idx]
                if isinstance(qty_val, (int, float)):
                    part["quantity"] = int(qty_val)
                else:
                    qty_str = str(qty_val).strip().replace(',', '')
                    part["quantity"] = int(float(qty_str)) if qty_str else 1
            except (ValueError, TypeError):
                part["quantity"] = 1

        if unit_idx is not None and unit_idx < len(row) and row[unit_idx]:
            part["unit"] = str(row[unit_idx]).strip()

        parts.append(part)

    wb.close()
    return parts


def read_rfq_csv(filepath: str) -> List[Dict]:
    """
    Read an RFQ CSV file and extract part information.
    Flexible column detection - looks for common column names.
    """
    parts = []

    # Common column name variations
    part_number_cols = ['part_number', 'part number', 'partnumber', 'pn', 'part#',
                        'part_no', 'part no', 'partno', 'item_number', 'item number',
                        'sku', 'mpn', 'manufacturer_part_number', 'nsn', 'niin']
    description_cols = ['description', 'desc', 'item_description', 'part_description',
                        'name', 'item_name', 'part_name', 'title']
    quantity_cols = ['quantity', 'qty', 'amount', 'count', 'units', 'order_qty']
    unit_cols = ['unit', 'uom', 'unit_of_measure', 'each']

    def find_column(headers: List[str], possible_names: List[str]) -> Optional[int]:
        """Find column index by trying multiple possible names."""
        headers_lower = [h.lower().strip() for h in headers]
        for name in possible_names:
            if name.lower() in headers_lower:
                return headers_lower.index(name.lower())
        return None

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        # Try to detect delimiter
        sample = f.read(2048)
        f.seek(0)

        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        reader = csv.reader(f, dialect)

        headers = next(reader)

        # Find column indices
        pn_idx = find_column(headers, part_number_cols)
        desc_idx = find_column(headers, description_cols)
        qty_idx = find_column(headers, quantity_cols)
        unit_idx = find_column(headers, unit_cols)

        if pn_idx is None:
            # If no part number column found, assume first column
            print("Warning: No part number column detected, using first column")
            pn_idx = 0

        print(f"Detected columns:")
        print(f"  Part Number: {headers[pn_idx] if pn_idx is not None else 'Column 0'}")
        print(f"  Description: {headers[desc_idx] if desc_idx is not None else 'Not found'}")
        print(f"  Quantity: {headers[qty_idx] if qty_idx is not None else 'Not found (defaulting to 1)'}")

        for row_num, row in enumerate(reader, start=2):
            if not row or not row[pn_idx].strip():
                continue

            part = {
                "part_number": row[pn_idx].strip(),
                "description": row[desc_idx].strip() if desc_idx is not None and desc_idx < len(row) else "",
                "quantity": 1,
                "unit": ""
            }

            # Parse quantity
            if qty_idx is not None and qty_idx < len(row):
                try:
                    qty_str = row[qty_idx].strip().replace(',', '')
                    part["quantity"] = int(float(qty_str)) if qty_str else 1
                except ValueError:
                    part["quantity"] = 1

            if unit_idx is not None and unit_idx < len(row):
                part["unit"] = row[unit_idx].strip()

            parts.append(part)

    return parts


def generate_report(parts_results: List[Dict], output_path: str):
    """Generate a detailed text report of supplier findings."""

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append("=" * 80)
    lines.append("RFQ SUPPLIER ANALYSIS REPORT")
    lines.append(f"Generated: {timestamp}")
    lines.append("=" * 80)
    lines.append("")

    # Summary section
    total_parts = len(parts_results)
    parts_with_suppliers = sum(1 for p in parts_results if p.get('suppliers'))
    exact_matches = sum(1 for p in parts_results if p.get('exact_match_found'))

    lines.append("SUMMARY")
    lines.append("-" * 40)
    lines.append(f"Total Parts Searched:     {total_parts}")
    lines.append(f"Parts with Suppliers:     {parts_with_suppliers}")
    lines.append(f"Exact Part Matches:       {exact_matches}")
    lines.append(f"No Suppliers Found:       {total_parts - parts_with_suppliers}")
    lines.append("")
    lines.append("")

    # Detailed results for each part
    for i, result in enumerate(parts_results, 1):
        lines.append("=" * 80)
        lines.append(f"PART {i} of {total_parts}")
        lines.append("=" * 80)
        lines.append("")
        lines.append(f"Part Number:     {result.get('part_number', 'N/A')}")
        lines.append(f"Description:     {result.get('description', 'N/A') or 'Not provided'}")
        lines.append(f"Quantity:        {result.get('quantity_requested', 1)}")
        lines.append(f"Exact Match:     {'YES' if result.get('exact_match_found') else 'NO'}")
        lines.append(f"Sources Found:   {result.get('sources_searched', 0)}")
        lines.append("")

        suppliers = result.get('suppliers', [])

        if suppliers:
            lines.append(f"SUPPLIERS FOUND: {len(suppliers)}")
            lines.append("-" * 40)

            for j, supplier in enumerate(suppliers, 1):
                lines.append("")
                lines.append(f"  Supplier {j}: {supplier.get('supplier_name', 'Unknown')}")
                lines.append(f"  Website:    {supplier.get('website', 'N/A')}")
                lines.append(f"  Confidence: {supplier.get('confidence', 'N/A')}")

                # Contact info
                contact = supplier.get('contact_info', {})
                if contact:
                    lines.append("")
                    lines.append("  Contact Information:")
                    if contact.get('phone'):
                        lines.append(f"    Phone:   {contact['phone']}")
                    if contact.get('email'):
                        lines.append(f"    Email:   {contact['email']}")
                    if contact.get('address'):
                        lines.append(f"    Address: {contact['address']}")

                # Pricing
                pricing = supplier.get('pricing', {})
                if pricing:
                    lines.append("")
                    lines.append("  Pricing Information:")
                    lines.append(f"    Unit Price:    {pricing.get('unit_price', 'Not Listed')}")
                    if pricing.get('currency') and pricing.get('currency') != 'USD':
                        lines.append(f"    Currency:      {pricing['currency']}")
                    if pricing.get('minimum_order_qty'):
                        lines.append(f"    Min Order Qty: {pricing['minimum_order_qty']}")
                    if pricing.get('lead_time'):
                        lines.append(f"    Lead Time:     {pricing['lead_time']}")
                    if pricing.get('price_notes'):
                        lines.append(f"    Notes:         {pricing['price_notes']}")

                # Ordering process
                if supplier.get('ordering_process'):
                    lines.append("")
                    lines.append("  How to Order:")
                    # Wrap long text
                    order_text = supplier['ordering_process']
                    wrapped = [order_text[i:i + 60] for i in range(0, len(order_text), 60)]
                    for line in wrapped:
                        lines.append(f"    {line}")

                # Additional notes
                if supplier.get('notes'):
                    lines.append("")
                    lines.append(f"  Notes: {supplier['notes']}")

                lines.append("")
                lines.append("  " + "-" * 36)

        else:
            lines.append("NO SUPPLIERS FOUND")
            lines.append("-" * 40)

        # Alternative sources and recommendations
        if result.get('alternative_sources'):
            lines.append("")
            lines.append("Alternative Sources:")
            lines.append(f"  {result['alternative_sources']}")

        if result.get('search_recommendations'):
            lines.append("")
            lines.append("Search Recommendations:")
            lines.append(f"  {result['search_recommendations']}")

        lines.append("")
        lines.append("")

    # Write report
    report_content = "\n".join(lines)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(report_content)

    return report_content


def main():
    parser = argparse.ArgumentParser(
        description="Find suppliers and pricing for parts in an RFQ file (CSV or Excel)"
    )
    parser.add_argument(
        "input_csv",
        metavar="INPUT_FILE",
        help="Path to the RFQ file (CSV, XLSX, or XLS) containing part numbers"
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output file path (default: rfq_results_TIMESTAMP.txt)"
    )
    parser.add_argument(
        "--delay", "-d",
        type=float,
        default=1.0,
        help="Delay between API requests in seconds (default: 1.0)"
    )
    parser.add_argument(
        "--limit", "-l",
        type=int,
        default=None,
        help="Limit number of parts to process (for testing)"
    )

    args = parser.parse_args()

    # Validate input file
    if not os.path.exists(args.input_csv):
        print(f"ERROR: Input file not found: {args.input_csv}")
        sys.exit(1)

    # Set output path
    if args.output:
        output_path = args.output
    else:
        # Create outputs folder if it doesn't exist
        outputs_dir = Path("outputs")
        outputs_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = outputs_dir / f"rfq_results_{timestamp}.txt"

    print("=" * 60)
    print("RFQ SUPPLIER FINDER")
    print("=" * 60)
    print(f"Input file:  {args.input_csv}")
    print(f"Output file: {output_path}")
    print(f"API delay:   {args.delay}s")
    print("")

    # Read the RFQ file (CSV or Excel)
    print("Reading RFQ file...")
    parts = read_rfq_file(args.input_csv)

    if not parts:
        print("ERROR: No parts found in the file")
        sys.exit(1)

    print(f"Found {len(parts)} parts to process")

    # Apply limit if specified
    if args.limit:
        parts = parts[:args.limit]
        print(f"Limited to first {args.limit} parts")

    print("")

    # Initialize supplier finder
    finder = SupplierFinder(delay=args.delay)

    # Process each part
    results = []
    for i, part in enumerate(parts, 1):
        print(f"\n[{i}/{len(parts)}] Processing: {part['part_number']}")

        result = finder.find_suppliers_for_part(
            part_number=part['part_number'],
            description=part.get('description', ''),
            quantity=part.get('quantity', 1)
        )

        results.append(result)

        # Show quick summary
        if result.get('suppliers'):
            print(f"    ✓ Found {len(result['suppliers'])} supplier(s)")
            if result.get('exact_match_found'):
                print(f"    ✓ Exact part match confirmed")
        else:
            print(f"    ✗ No suppliers found")

    # Generate report
    print("\n" + "=" * 60)
    print("Generating report...")

    report = generate_report(results, output_path)

    print(f"\n✓ Report saved to: {output_path}")

    # Print summary
    total = len(results)
    with_suppliers = sum(1 for r in results if r.get('suppliers'))
    exact = sum(1 for r in results if r.get('exact_match_found'))

    print("")
    print("SUMMARY")
    print("-" * 30)
    print(f"Total parts:           {total}")
    print(f"Parts with suppliers:  {with_suppliers}")
    print(f"Exact matches:         {exact}")
    print(f"No suppliers found:    {total - with_suppliers}")


if __name__ == "__main__":
    main()